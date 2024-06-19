import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import re

# Function to filter out non-English characters
def filter_non_english_characters(text):
    if text is not None:
        return re.sub(r'[^\x00-\x7F]+', '', text) # Remove non-ASCII characters
    return text

# Load the workbook
workbook_path = "output.xlsx"
dataframe = openpyxl.load_workbook(workbook_path)

# Select the active sheet
dataframe1 = dataframe.active

# Modify the workbook
for row in dataframe1.iter_rows():
    for cell in row:
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.value = filter_non_english_characters(cell.value)

# Try saving the modified workbook to a new file to preserve the original
new_workbook_path = "ModifiedFBS_NOVO_DS1_R6-2_240408.xlsx"
dataframe.save(new_workbook_path)

# Try opening the new workbook to check for issues
try:
    test_open = openpyxl.load_workbook(new_workbook_path)
    print("Workbook saved and verified successfully.")
except InvalidFileException as e:
    print(f"Failed to verify the workbook: {e}")
