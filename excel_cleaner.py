import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import re
import os
from tkinter import simpledialog
import tkinter as tk

#Function to filter out non-English characters
def filter_non_english_characters(text):
    if text is not None:
        return re.sub(r'[^\x00-\x7F]+', '', text) # Remove non-ASCII characters
    return text

# change the character ö, å , ä to oe, aa, ae 
def change_swedish_characters(text):
    # use regex to replace the characters
    if text is not None:
        text = re.sub(r'ö', 'oe', text)
        text = re.sub(r'å', 'aa', text)
        text = re.sub(r'ä', 'ae', text)
    return text

# change Ø to phase 
def change_phi(text):
    # use regex to replace the characters
    if text is not None:
        text = re.sub(r'Ø', 'phase', text)
    return text

def change_degrees(text):
    # use regex to replace the characters
    if text is not None:
        text = re.sub(r'(\d+)º', r'\1 degrees', text)
    return text

def replace_non_ascii_characters(text):
    if text is not None:
        # Replace the ≤ character with <=
        text = re.sub(r'≤', '<=', text)
        # Add other replacements here if needed
    return text

def read_files():
    # read all excel files in the current directory
    file_list = [file for file in os.listdir() if file.endswith(".xlsx")]
    return file_list

# Load the workbook
def clean_excel():
    # chose the file from the list
    root = tk.Tk()
    root.withdraw()
    file_list = read_files()
    if not file_list:
        simpledialog.messagebox.showinfo("Info", "No Excel files found in the current directory.")
        exit()
    selected_file = simpledialog.askinteger("Input", f"Select the file to clean by entering a number from 1 to {len(file_list)} from below list:\n\n" + "\n".join(f"{i+1}. {file}" for i, file in enumerate(file_list)), parent=root, minvalue=1, maxvalue=len(file_list))
    workbook_path = file_list[selected_file - 1]
    dataframe = openpyxl.load_workbook(workbook_path)

    # Select the active sheet
    dataframe1 = dataframe.active

    # Modify the workbook
    for row in dataframe1.iter_rows():
        for cell in row:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell) and cell.value is not None:
                cell.value = filter_non_english_characters(cell.value)
                ################## delet the comment below to use the functions #########################
                # cell.value = change_phi(cell.value)
                # cell.value = change_swedish_characters(cell.value)
                # cell.value = change_degrees(cell.value)
                # cell.value = replace_non_ascii_characters(cell.value)

    # Try saving the modified workbook to a new file to preserve the original. Orginal file + Modified
    new_workbook_path = f"{os.path.splitext(workbook_path)[0]}_cleaned.xlsx"
    dataframe.save(new_workbook_path)

    # Try opening the new workbook to check for issues
    try:
        test_open = openpyxl.load_workbook(new_workbook_path)
        print("Workbook saved and verified successfully.")
    except InvalidFileException as e:
        print(f"Failed to verify the workbook: {e}")

clean_excel()
