import tkinter as tk
from tkinter import simpledialog, messagebox
import os
import openpyxl


def copy_values_to_template(source_file, template_file, output_file, max_column=17, max_row=100):
    # Load the workbook from the source file and the template file
    source_wb = openpyxl.load_workbook(source_file)
    template_wb = openpyxl.load_workbook(template_file)
    # Get the active sheet from the source and template workbooks
    source_ws = source_wb.active
    template_ws = template_wb.active
    # Iterate over the rows and columns within the defined range
    for row in source_ws.iter_rows(min_row=1, max_col=max_column, max_row=max_row):
        for cell in row:
            # Copy the value from the source cell to the corresponding cell in the template
            template_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    # Save the modified template to a new output file
    template_wb.save(output_file)



def get_column_row():
    root = tk.Tk()
    root.withdraw()  # Hide the tkinter root window
    max_column = simpledialog.askinteger("Input", "Enter the maximum column to copy (e.g., 17 for column Q):", parent=root, minvalue=1, maxvalue=256)
    max_row = simpledialog.askinteger("Input", "Enter the maximum row to copy (e.g., 100):", parent=root, minvalue=1, maxvalue=1048576)
    root.destroy()
    return max_column, max_row

def read_files():
    # read all excel files in the current directory
    file_list = [file for file in os.listdir() if file.endswith(".xlsx")]
    return file_list


def select_files(file_list):
    # Let the user select the source and template files
    root = tk.Tk()
    root.withdraw()
    ## show box of all files

    max_files = len(file_list)
    # shose the numberr from the list 
    source_file_number = simpledialog.askinteger("Input", f"Select the source file by entering a number from 1 to {max_files} from below list:\n\n" + "\n".join(f"{i+1}. {file}" for i, file in enumerate(file_list)), parent=root, minvalue=1, maxvalue=max_files)
    # delete the source file from the list
    template_file_number = simpledialog.askinteger("Input", f"Select the template file by entering a number from 1 to {max_files} from below list:\n\n" + "\n".join(f"{i+1}. {file}" for i, file in enumerate(file_list)), parent=root, minvalue=1, maxvalue=max_files)
    root.destroy()

    if source_file_number is not None:
        source_file = file_list[source_file_number - 1]
        if template_file_number is not None:
            template_file = file_list[template_file_number - 1]
            return source_file, template_file
    return None

# Usage of the function
 
if __name__ == '__main__':
    file_list = read_files()
    if not file_list:
        messagebox.showinfo("Files List", "No Excel files found in the current directory.")
        exit()
    source_file, template_file = select_files(file_list)
    max_column, max_row = get_column_row()
    # Ask the user for the output file name
    input = simpledialog.askstring("Input", "Enter the name of the output file (e.g., output):")
    file_name = input + ".xlsx" if input else "output.xlsx"
    output_file = file_name
    copy_values_to_template(source_file, template_file, output_file, max_column, max_row)
